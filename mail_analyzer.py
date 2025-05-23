import pandas as pd
import time
from difflib import get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Couleurs ANSI pour l'affichage
GREEN = '\033[92m'
YELLOW = '\033[93m'
RED = '\033[91m'
RESET = '\033[0m'

# Constantes
INVALID_CHARS = [',', ';', ':', '!']
VALID_EXTENSIONS = ['com', 'fr', 'net', 'org', 'mc', 'es', 'io', 'uk', 'be', 'eu', 'pro', 'no', 'pt', 'de', 'au', 'ch', 'art', 'bio', 'wine', 'solar', 'edu', 'pw', 'info']

def afficher_message_accueil():
    """
    Affiche un message d'accueil pour l'utilisateur.
    """
    print(f"{YELLOW}Bienvenue dans le programme d'analyse d'emails !{RESET}")
    print(f"{YELLOW}Pensez à placer le fichier excel (.xlsx) à analyser dans le même dossier que ce script{RESET}")

def charger_fichier_excel():
    """
    Charge un fichier Excel à partir du nom fourni par l'utilisateur.
    Retourne un DataFrame pandas contenant les données du fichier.
    """
    while True:
        try:
            ExcelName = input("\nVeuillez entrer le nom du fichier: ")
            if not ExcelName.endswith('.xlsx'):
                ExcelName += '.xlsx'
            df = pd.read_excel(ExcelName)
            print(f"{GREEN}Chargement du fichier Excel {ExcelName}...{RESET}\n")
            return df
        except FileNotFoundError:
            print(f"{RED}Fichier introuvable. Veuillez vérifier le nom du fichier et réessayer.{RESET}")

def obtenir_colonne_email(df):
    """
    Demande à l'utilisateur de spécifier la colonne contenant les adresses email.
    Retourne le nom de la colonne sélectionnée.
    """
    while True:
        column_names = df.columns
        print(f"{YELLOW}Colonnes disponibles dans le fichier :{RESET}")
        print(column_names)
        email_column = input(f"{YELLOW}Entrez le nom de la colonne contenant les adresses email: {RESET}")
        if email_column in column_names:
            return email_column
        else:
            print(f"{RED}Nom de colonne invalide. Veuillez réessayer.{RESET}")

def analyser_emails(emails):
    """
    Analyse une série d'emails pour extraire les extensions, domaines, emails invalides et multiples.
    Retourne quatre listes : extensions, domaines, emails_invalides, emails_multiples.
    """
    emails_valides = []
    emails_invalides = []
    emails_multiples = []

    print(f"{GREEN}Analyse des adresses email en cours...{RESET}")

    for index, email_entry in enumerate(emails):
        email_entry = email_entry.strip()
        email = email_entry.split(';')[0]
        split_email = email_entry.split(';')

        if len(split_email) > 1:
            serie_email = ';'.join(split_email[0:]).strip()
            emails_multiples.append((index + 2, serie_email))

        if any(char in email for char in INVALID_CHARS):
            emails_invalides.append((index + 2, email))
            continue

        if '@' in email and '.' in email:
            extension = email.split('.')[-1]
            if len(extension) >= 2 and len(extension) <= 7:
                if extension not in VALID_EXTENSIONS:
                    emails_invalides.append((index + 2, email))
                else:
                    emails_valides.append((index + 2, email))
        else:
            emails_invalides.append((index + 2, email))

    return emails_valides, emails_invalides, emails_multiples

def afficher_resultats(emails_valides, emails_invalides, emails_multiples):
    """
    Affiche les résultats de l'analyse des emails.
    """
    print(f"\n{GREEN}Analyse terminée. Voici les résultats :{RESET}")
    time.sleep(1)

    print(f"\n{YELLOW}Emails valides :{RESET}")
    counter = 0
    for index, email in emails_valides:
        counter += 1
        print(f"Index {index}: {email}")
    print(f"\n{GREEN}Nombre d'emails valides : {counter}{RESET}")
    if not emails_invalides and not emails_multiples:
        print(f"{GREEN}Tous les emails ont été détectés valides.{RESET}")
    else:
        if emails_invalides:
            print(f"\n{YELLOW}Emails invalides :{RESET}")
            for index, email in emails_invalides:
                print(f"Index {index}: {email}")
        if emails_multiples:
            print(f"\n{YELLOW}Emails multiples :{RESET}")
            for index, email in emails_multiples:
                print(f"Index {index}: {email}")

def correct_extension(email):
    """
    Corrige l'extension d'un email.
    """
    if '.' in email:
        parts = email.split('.')
        extension = parts[-1]

        if len(extension) > 3 and not extension.isalpha():
            extension = extension.rstrip(''.join(c for c in extension if not c.isalpha()))
            email = '.'.join(parts[:-1] + [extension])

        if extension not in VALID_EXTENSIONS:
            similar_extensions = get_close_matches(extension, VALID_EXTENSIONS, n=1, cutoff=0.6)
            if similar_extensions:
                email = '.'.join(parts[:-1] + [similar_extensions[0]])

    if '@' in email and '.' not in email.split('@')[-1]:
        domain_part = email.split('@')[-1]
        for ext in VALID_EXTENSIONS:
            if ext in domain_part:
                email = email.replace(domain_part, domain_part.replace(ext, f'.{ext}'))
                break

    return email

def correct_and_update_email(df_corrected, email_column, emails_valides, index, email, is_multiple=False):
    """
    Corrige un email et met à jour le DataFrame et la liste des emails valides.
    """
    if is_multiple:
        first_email = email.split(';')[0].strip()
        corrected_email = first_email.lower()
    else:
        corrected_email = email.lower()

    corrected_email = corrected_email.replace(',', '.')
    corrected_email = ''.join(char for char in corrected_email if char not in INVALID_CHARS)
    final_email = correct_extension(corrected_email)

    if final_email.endswith('.'):
        final_email = final_email.rstrip('.')

    if not is_multiple and '@' in final_email and '.' not in final_email.split('@')[-1]:
        domain_part = final_email.split('@')[-1]
        print(f"\n{RED}Extension manquante pour l'email: {final_email}{RESET}")
        user_extension = input(f"{YELLOW}Veuillez entrer une extension valide pour cet email (par exemple, com, fr, net) : {RESET}\n")
        if user_extension in VALID_EXTENSIONS:
            final_email = final_email.replace(domain_part, f"{domain_part}.{user_extension}")

    df_corrected.at[index - 2, email_column] = final_email
    emails_valides.append((index, final_email))

def revise_file(df, email_column, emails_valides, emails_invalides, emails_multiples):
    """
    Corrige le fichier Excel en réécrivant les emails invalides et multiples.
    """
    df_corrected = df.copy()

    for index, email in emails_invalides:
        correct_and_update_email(df_corrected, email_column, emails_valides, index, email, is_multiple=False)

    for index, email in emails_multiples:
        correct_and_update_email(df_corrected, email_column, emails_valides, index, email, is_multiple=True)

    validation = input(f"\n{YELLOW}Voulez-vous enregistrer le fichier corrigé? (oui/non) : {RESET}")

    if validation.lower() == 'oui':
        output_file = "corrected_file.xlsx"
        df_corrected.to_excel(output_file, index=False)
        print(f"{GREEN}Fichier corrigé enregistré sous {output_file}.{RESET}")
    else:
        print(f"{RED}Enregistrement du fichier corrigé annulé.{RESET}")

    return df_corrected

def enregistrer_rapport(df, emails_invalides, emails_multiples):
    """
    Enregistre les résultats de l'analyse dans un fichier Excel et un fichier texte.
    """
    validation = input(f"\n{YELLOW}Voulez-vous enregistrer un nouveau fichier avec les données invalides surlignées? (oui/non) : {RESET}")

    if validation.lower() == 'oui':
        output_file = 'colored_data.xlsx'
        df.to_excel(output_file, index=False)

        book = load_workbook(output_file)
        ws = book.active
        fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

        all_emails_to_highlight = emails_invalides + emails_multiples
        for index, _ in all_emails_to_highlight:
            row = ws[index]
            for cell in row:
                cell.fill = fill

        book.save(output_file)
        print(f"{GREEN}Données enregistrées avec succès sous {output_file}.{RESET}")

        with open('rapport_emails.txt', 'w', encoding='utf-8') as file:
            file.write("Ce fichier contient un rapport des emails détectés invalides ou multiples. Avec 'Index' correspondant à la ligne Excel, et l'email associé à cette ligne. \n\n")
            file.write("Emails invalides :\n")
            for index, email in emails_invalides:
                file.write(f"Index {index}: {email}\n")
            file.write("\nEmails multiples :\n")
            for index, email in emails_multiples:
                file.write(f"Index {index}: {email}\n")

        print(f"{GREEN}Le rapport a été écrit dans le fichier rapport_emails.txt.{RESET}")
    else:
        print(f"{RED}Enregistrement annulé.{RESET}")

def main():
    """
    Fonction principale qui orchestre l'exécution du programme.
    """
    afficher_message_accueil()
    df = charger_fichier_excel()
    email_column = obtenir_colonne_email(df)
    emails = df[email_column]
    emails_valides, emails_invalides, emails_multiples = analyser_emails(emails)
    afficher_resultats(emails_valides, emails_invalides, emails_multiples)
    df_corrected = revise_file(df, email_column, emails_valides, emails_invalides, emails_multiples)
    enregistrer_rapport(df, emails_invalides, emails_multiples)
    print(f"{GREEN}Fermeture du programme...{RESET}")
    time.sleep(2)

if __name__ == "__main__":
    main()
